import openpyxl

class Mapp():
    
    def __init__(self,fichier_map,fichier_wave,feuille):
        
        self.mapp = self.lecture_fichier_color(fichier_map,feuille)
        self.wave = self.lecture_fichier_wave(fichier_wave,feuille)
    
    def lecture_fichier_color(self,fichier,feuille):
        # Ouvrir le fichier Excel
        wb = openpyxl.load_workbook(fichier)
        sheet = wb[feuille]
        
        color_matrix = []
        
        # Parcourir les cellules
        for row in sheet.iter_rows():
            color_row = []
            i = 0
            #print(row)
            fill0 = row[0].fill
            
            if fill0.fgColor.rgb == 'FFC00000':
                        
                wb.close()
                return color_matrix
            
            else:
                while i != len(row):
                    cell = row[i]
                    i +=1
                    fill = cell.fill
                    couleur = fill.fgColor.rgb
                    if couleur == 'FFFFFFFF' or couleur == '00000000':
                        color_row.append('Chemin')
                    
                    elif couleur == 'FFC00000':
                        i = len(row)  
                    
                    elif couleur == 'FF4D4D4D':
                        color_row.append('Montagne')
                        
                    elif couleur == 'FF0066FF':
                        color_row.append('Riviere')
                        
                    elif couleur == 'FF006600':
                        color_row.append('Foret')
                    
                    elif couleur == 'FF996633':
                        color_row.append('Pont')
                        
                    elif couleur == 'FFD60093':
                        color_row.append('Depart')
                        
                    elif couleur == 'FFFF3300':
                        color_row.append('Arrivee')
                    else :
                        color_row.append(couleur)
                        
                color_matrix.append(color_row)
                
    def lecture_fichier_wave(self,fichier,feuille): #renvoie un dico qui pour chaque wave sous forme de numéro lui associe une liste qui représente chaque tic de la manche
        
        wb = openpyxl.load_workbook(fichier)
        sheet = wb[feuille]
        
        wave_ennemi_dico = {}
        liste = []
        str_ennemi = ''
        ennemi_type =''
        i = 0
        for row in sheet.iter_rows(values_only=True):
            val_0 = str(row[0])
            i = 0
            if val_0[:4] != "wave" :
                while i < len(row) and row[i] != None :
                    valeur = str(row[i])
                    if valeur != '0':
                        if i == 0:
                            ennemi_type = 'n'
                        elif i == 1:
                            ennemi_type = 'b'
                        str_ennemi += ennemi_type
                    i += 1
                
                liste.append(str_ennemi)
                str_ennemi = ''
                
            else :
                wave_ennemi_dico[row[0][4:]] = liste
                liste = []
        
        return wave_ennemi_dico
    
    def affichage_map(self):
        for row in self.map:
            print(row)
            

    def ajouter_tourelle(self,coor,nom):
        liste = self.mapp[coor[0]]
        liste[coor[1]] = nom

coucou = Mapp("Map.xlsx","Wave.xlsx","Feuil3")
print(coucou.mapp)
coucou.ajouter_tourelle([0,0],'tourelle')
print(coucou.mapp)